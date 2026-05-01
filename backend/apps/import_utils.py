from __future__ import annotations

import csv
import io
import json

_MAX_BYTES = 5 * 1024 * 1024  # 5 MB


class ImportFileError(ValueError):
    pass


def parse_import_file(file_obj, filename: str) -> list[dict]:
    """Parse a CSV, JSON, or XLSX upload into a list of lowercase-keyed string dicts.

    Raises ImportFileError with a human-readable message on failure.
    Accepted extensions: .csv  .json  .xlsx  .xls
    """
    if file_obj.size > _MAX_BYTES:
        raise ImportFileError("File too large (max 5 MB).")

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "csv"

    # ── CSV ──────────────────────────────────────────────────────────────────
    if ext == "csv":
        decoded = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        if not reader.fieldnames:
            raise ImportFileError("CSV has no headers.")
        return [{k.strip().lower(): (v or "").strip() for k, v in row.items()} for row in reader]

    # ── JSON ─────────────────────────────────────────────────────────────────
    if ext == "json":
        try:
            data = json.loads(file_obj.read().decode("utf-8"))
        except json.JSONDecodeError as exc:
            raise ImportFileError(f"Invalid JSON: {exc}") from exc
        if not isinstance(data, list):
            raise ImportFileError("JSON root must be an array of objects.")
        rows: list[dict] = []
        for i, item in enumerate(data):
            if not isinstance(item, dict):
                raise ImportFileError(f"JSON item at index {i} must be an object, not {type(item).__name__}.")
            rows.append({k.strip().lower(): str(v or "").strip() for k, v in item.items()})
        return rows

    # ── XLSX / XLS ────────────────────────────────────────────────────────────
    if ext in ("xlsx", "xls"):
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError as exc:
            raise ImportFileError("openpyxl is not installed — cannot read .xlsx files.") from exc
        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
            ws = wb.active
            iter_rows = ws.iter_rows(values_only=True)
            raw_headers = next(iter_rows, None)
            if not raw_headers:
                raise ImportFileError("Excel file has no header row.")
            headers = [str(h).strip().lower() if h is not None else "" for h in raw_headers]
            rows = []
            for row in iter_rows:
                d = {headers[i]: str(row[i] or "").strip() for i in range(len(headers)) if i < len(row)}
                rows.append(d)
            return rows
        except ImportFileError:
            raise
        except Exception as exc:
            raise ImportFileError(f"Could not read Excel file: {exc}") from exc

    raise ImportFileError(f"Unsupported format '.{ext}'. Accepted: .csv  .json  .xlsx")
