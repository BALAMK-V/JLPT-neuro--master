from __future__ import annotations

import csv
import io
import json
import os
import re
import sqlite3
import tempfile
import zipfile
from dataclasses import dataclass

from django.db import transaction
from django.db.utils import IntegrityError
from django.db.models import Count, Q
from django.utils import timezone
from rest_framework import permissions, status, viewsets
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.users.permissions import IsManagementUser

from apps.content.models import Kanji, JLPTLevel, Vocabulary
from apps.users.models import UserProfile

from .models import Card, Deck, ImportLog
from .serializers import CardSerializer, DeckSerializer, ImportLogSerializer, ReviewApplySerializer


def _ensure_default_decks(user) -> tuple[Deck, Deck]:
    """Create 2 locked decks per user: Kanji + Vocabulary."""

    UserProfile.objects.get_or_create(user=user)
    level = getattr(user.profile, "jlpt_level", JLPTLevel.N2)

    def get_or_create_locked(system_key: str, name: str, deck_type: str) -> Deck:
        try:
            deck, _created = Deck.objects.get_or_create(
                user=user,
                system_key=system_key,
                defaults={
                    "name": name,
                    "deck_type": deck_type,
                    "jlpt_level": level,
                    "is_locked": True,
                },
            )
        except IntegrityError:
            # If user already created a deck with the same name, reuse it but lock it.
            deck = Deck.objects.get(user=user, name=name)
            deck.system_key = system_key
            deck.deck_type = deck_type
            deck.is_locked = True

        # Keep default deck level in sync with profile level.
        changed = False
        if deck.jlpt_level != level:
            deck.jlpt_level = level
            changed = True
        if deck.name != name:
            deck.name = name
            changed = True
        if deck.deck_type != deck_type:
            deck.deck_type = deck_type
            changed = True
        if not deck.is_locked:
            deck.is_locked = True
            changed = True
        if deck.system_key != system_key:
            deck.system_key = system_key
            changed = True
        if changed:
            deck.save()
        return deck

    kanji_deck = get_or_create_locked("kanji_default", "Kanji (Default)", Deck.DeckType.KANJI)
    vocab_deck = get_or_create_locked("vocab_default", "Vocabulary (Default)", Deck.DeckType.VOCAB)
    return kanji_deck, vocab_deck


def _sync_locked_deck(deck: Deck) -> int:
    """Populate locked default decks with content items for the deck JLPT level."""

    if not deck.is_locked:
        return 0

    level = deck.jlpt_level or JLPTLevel.N2
    created = 0

    if deck.deck_type == Deck.DeckType.KANJI:
        # Remove cards that no longer match the deck level.
        Card.objects.filter(deck=deck, kanji__isnull=False).exclude(kanji__jlpt_level=level).delete()
        existing = set(Card.objects.filter(deck=deck, kanji__isnull=False).values_list("kanji_id", flat=True))
        missing = list(Kanji.objects.filter(jlpt_level=level).exclude(id__in=existing))
        cards = [
            Card(
                deck=deck,
                kanji=k,
                front=k.character,
                back=f"{k.meaning_en}\nonyomi: {k.onyomi}\nkunyomi: {k.kunyomi}".strip(),
                furigana=k.onyomi,
                tags=[level, "kanji"],
            )
            for k in missing
        ]
        if cards:
            Card.objects.bulk_create(cards, batch_size=500, ignore_conflicts=True)
            created = len(cards)
        return created

    if deck.deck_type == Deck.DeckType.VOCAB:
        Card.objects.filter(deck=deck, vocab__isnull=False).exclude(vocab__jlpt_level=level).delete()
        existing = set(Card.objects.filter(deck=deck, vocab__isnull=False).values_list("vocab_id", flat=True))
        missing = list(Vocabulary.objects.filter(jlpt_level=level).exclude(id__in=existing))
        cards = [
            Card(
                deck=deck,
                vocab=v,
                front=v.word,
                back=f"{v.meaning_en}\nreading: {v.reading}".strip(),
                furigana=v.reading,
                tags=[level, "vocab"],
            )
            for v in missing
        ]
        if cards:
            Card.objects.bulk_create(cards, batch_size=500, ignore_conflicts=True)
            created = len(cards)
        return created

    return 0


class DeckViewSet(viewsets.ModelViewSet):
    serializer_class = DeckSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ["deck_type", "jlpt_level"]
    search_fields = ["name"]
    ordering_fields = ["id", "updated_at", "name"]

    def get_queryset(self):  # type: ignore[no-untyped-def]
        _ensure_default_decks(self.request.user)
        now = timezone.now()
        return (
            Deck.objects.filter(user=self.request.user)
            .annotate(
                total_cards=Count("cards"),
                due_count=Count("cards", filter=Q(cards__suspended=False, cards__due_at__lte=now)),
            )
            .order_by("-updated_at")
        )

    def perform_create(self, serializer):  # type: ignore[no-untyped-def]
        serializer.save(user=self.request.user, is_locked=False, system_key="")

    def update(self, request, *args, **kwargs):  # type: ignore[no-untyped-def]
        obj = self.get_object()
        if obj.is_locked:
            return Response({"detail": "This deck is locked."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):  # type: ignore[no-untyped-def]
        obj = self.get_object()
        if obj.is_locked:
            return Response({"detail": "This deck is locked."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):  # type: ignore[no-untyped-def]
        obj = self.get_object()
        if obj.is_locked:
            return Response({"detail": "This deck is locked."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


class CardViewSet(viewsets.ModelViewSet):
    serializer_class = CardSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ["deck", "suspended", "kanji", "vocab"]
    search_fields = ["front", "back"]
    ordering_fields = ["id", "due_at", "updated_at"]

    def get_queryset(self):  # type: ignore[no-untyped-def]
        return Card.objects.filter(deck__user=self.request.user).select_related("deck", "kanji", "vocab")

    def list(self, request, *args, **kwargs):  # type: ignore[no-untyped-def]
        deck_id = request.query_params.get("deck")
        if deck_id:
            deck = Deck.objects.filter(id=deck_id, user=request.user).first()
            if deck and deck.is_locked:
                _sync_locked_deck(deck)
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):  # type: ignore[no-untyped-def]
        deck_id = request.data.get("deck")
        if deck_id:
            deck = Deck.objects.filter(id=deck_id, user=request.user).first()
            if deck and deck.is_locked:
                return Response({"detail": "Cards in this deck are locked."}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):  # type: ignore[no-untyped-def]
        obj = self.get_object()
        if obj.deck.is_locked:
            return Response({"detail": "Cards in this deck are locked."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):  # type: ignore[no-untyped-def]
        obj = self.get_object()
        if obj.deck.is_locked:
            return Response({"detail": "Cards in this deck are locked."}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):  # type: ignore[no-untyped-def]
        obj = self.get_object()
        if obj.deck.is_locked:
            return Response({"detail": "Cards in this deck are locked."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


class FlashNextView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):  # type: ignore[no-untyped-def]
        deck_id = request.query_params.get("deck_id")
        limit = int(request.query_params.get("limit") or 20)
        limit = max(1, min(200, limit))

        qs = Card.objects.filter(deck__user=request.user, suspended=False, due_at__lte=timezone.now()).select_related("deck", "kanji", "vocab")
        if deck_id:
            deck = Deck.objects.filter(id=deck_id, user=request.user).first()
            if deck and deck.is_locked:
                _sync_locked_deck(deck)
            qs = qs.filter(deck_id=deck_id)

        items = list(qs.order_by("due_at", "id")[:limit])
        return Response({"count": len(items), "results": CardSerializer(items, many=True).data})


class FlashReviewView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):  # type: ignore[no-untyped-def]
        s = ReviewApplySerializer(data=request.data)
        s.is_valid(raise_exception=True)

        try:
            card = Card.objects.select_related("deck").get(id=s.validated_data["card_id"], deck__user=request.user)
        except Card.DoesNotExist:
            return Response({"detail": "Card not found."}, status=status.HTTP_404_NOT_FOUND)
        card.apply_rating(s.validated_data["rating"])
        card.save()
        return Response(CardSerializer(card).data, status=status.HTTP_200_OK)


@dataclass(frozen=True)
class ImportResult:
    created: int
    skipped: int


class FlashDueAllView(APIView):
    """GET /api/flash/due-all/ — total due count + first N cards across all decks."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):  # type: ignore[no-untyped-def]
        limit = int(request.query_params.get("limit") or 20)
        limit = max(1, min(200, limit))

        now = timezone.now()
        qs = Card.objects.filter(
            deck__user=request.user, suspended=False, due_at__lte=now
        ).select_related("deck", "kanji", "vocab")

        total_due = qs.count()
        items = list(qs.order_by("due_at", "id")[:limit])
        return Response(
            {"total_due": total_due, "count": len(items), "results": CardSerializer(items, many=True).data}
        )


class FlashLeechesView(APIView):
    """GET /api/flash/leeches/ — list suspended leech cards (lapses >= 8).
       POST /api/flash/leeches/<id>/unsuspend/ — manually unsuspend a leech."""

    LEECH_THRESHOLD = 8
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):  # type: ignore[no-untyped-def]
        leeches = Card.objects.filter(
            deck__user=request.user, suspended=True, lapses__gte=self.LEECH_THRESHOLD
        ).select_related("deck", "kanji", "vocab").order_by("-lapses")
        return Response({"count": leeches.count(), "results": CardSerializer(leeches, many=True).data})

    def post(self, request, card_id: int):  # type: ignore[no-untyped-def]
        try:
            card = Card.objects.get(id=card_id, deck__user=request.user)
        except Card.DoesNotExist:
            return Response({"detail": "Card not found."}, status=status.HTTP_404_NOT_FOUND)
        card.suspended = False
        card.lapses = 0
        card.repetitions = 0
        card.interval_days = 0
        card.due_at = timezone.now()
        card.save()
        return Response(CardSerializer(card).data)


_HTML_TAG = re.compile(r"<[^>]+>")


def _strip_html(text: str) -> str:
    return _HTML_TAG.sub("", text).strip()


def _rows_from_file(file_obj, ext: str) -> tuple[list[dict] | None, str | None]:
    """Parse an uploaded file into a list of lowercase-keyed dicts.

    Supported formats: csv, tsv, txt (Anki text export), json, xlsx, apkg.
    Returns (rows, error_message).
    """
    ext = ext.lower().lstrip(".")

    # ── CSV ──────────────────────────────────────────────────────────────────
    if ext == "csv":
        decoded = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        if not reader.fieldnames:
            return None, "CSV has no headers."
        rows = [{k.strip().lower(): (v or "").strip() for k, v in row.items()} for row in reader]
        return rows, None

    # ── TSV / TXT (Anki text export) ─────────────────────────────────────────
    if ext in ("tsv", "txt"):
        decoded = file_obj.read().decode("utf-8-sig")
        rows = []
        for line in decoded.splitlines():
            if not line.strip() or line.startswith("#"):
                continue
            parts = line.split("\t")
            if len(parts) < 2:
                continue
            tags_raw = parts[2].strip() if len(parts) > 2 else ""
            rows.append({
                "front": _strip_html(parts[0].strip()),
                "back": _strip_html(parts[1].strip()),
                "tags": ";".join(tags_raw.split()) if tags_raw else "",
            })
        return rows, None

    # ── JSON ─────────────────────────────────────────────────────────────────
    if ext == "json":
        try:
            data = json.loads(file_obj.read().decode("utf-8-sig"))
        except json.JSONDecodeError as e:
            return None, f"Invalid JSON: {e}"
        if not isinstance(data, list):
            return None, "JSON must be an array of card objects or arrays."
        rows = []
        for item in data:
            if isinstance(item, dict):
                rows.append({k.lower(): str(v or "").strip() for k, v in item.items()})
            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                tags_raw = str(item[2]).strip() if len(item) > 2 else ""
                rows.append({
                    "front": str(item[0]).strip(),
                    "back": str(item[1]).strip(),
                    "tags": ";".join(tags_raw.split()) if tags_raw else "",
                })
        return rows, None

    # ── XLSX ─────────────────────────────────────────────────────────────────
    if ext == "xlsx":
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError:
            return None, "openpyxl is not installed — cannot parse .xlsx files."
        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
            ws = wb.active
            iter_rows = ws.iter_rows(values_only=True)
            raw_headers = next(iter_rows, None)
            if not raw_headers:
                return None, "Excel file has no header row."
            headers = [str(h).strip().lower() if h is not None else "" for h in raw_headers]
            rows = []
            for row in iter_rows:
                d = {headers[i]: str(row[i] or "").strip() for i in range(len(headers)) if i < len(row)}
                rows.append(d)
            return rows, None
        except Exception as exc:
            return None, f"Could not read .xlsx file: {exc}"

    # ── APKG (Anki package) ───────────────────────────────────────────────────
    if ext == "apkg":
        try:
            with zipfile.ZipFile(file_obj, "r") as zf:
                db_name = next(
                    (n for n in zf.namelist() if n in ("collection.anki21", "collection.anki2")),
                    None,
                )
                if not db_name:
                    return None, "No Anki collection database found in .apkg file."
                with tempfile.NamedTemporaryFile(suffix=".anki2", delete=False) as tmp:
                    tmp.write(zf.read(db_name))
                    tmp_path = tmp.name

            try:
                conn = sqlite3.connect(tmp_path)
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("SELECT models FROM col LIMIT 1")
                col_row = cur.fetchone()
                models: dict = json.loads(col_row["models"]) if col_row else {}
                cur.execute("SELECT mid, flds, tags FROM notes")
                note_rows = cur.fetchall()
                conn.close()
            finally:
                os.unlink(tmp_path)

            _front_keys = {"front", "expression", "word", "kanji", "vocabulary", "vocab", "japanese", "question"}
            _back_keys = {"back", "meaning", "definition", "english", "answer", "reading", "translation"}

            rows = []
            for note in note_rows:
                mid = str(note["mid"])
                flds = note["flds"].split("\x1f")
                tags_raw = (note["tags"] or "").strip()
                tags = ";".join(tags_raw.split()) if tags_raw else ""

                model = models.get(mid, {})
                field_names = [f["name"].lower() for f in model.get("flds", [])]

                front = _strip_html(flds[0]) if flds else ""
                back = _strip_html(flds[1]) if len(flds) > 1 else ""

                if field_names:
                    field_dict = {name: _strip_html(flds[i]) for i, name in enumerate(field_names) if i < len(flds)}
                    for k in _front_keys:
                        if field_dict.get(k):
                            front = field_dict[k]
                            break
                    for k in _back_keys:
                        if field_dict.get(k):
                            back = field_dict[k]
                            break

                if front and back:
                    rows.append({"front": front, "back": back, "tags": tags})
            return rows, None

        except zipfile.BadZipFile:
            return None, "Invalid .apkg file (not a valid ZIP archive)."
        except Exception as exc:
            return None, f"Could not parse .apkg file: {exc}"

    return None, f"Unsupported file format: .{ext}. Supported: csv, tsv, txt, json, xlsx, apkg."


class FlashImportView(APIView):
    """Import flashcards from CSV, TSV/TXT, JSON, XLSX, or Anki APKG.

    Multipart POST:
      - import_file  (required; also accepts legacy field name csv_file)
      - deck_id (optional) OR deck_name + deck_type (optional)

    For CSV/XLSX, headers (case-insensitive):
      front, back, tags, kanji_character, vocab_word, vocab_reading, jlpt_level
    For TSV/TXT (Anki text export):
      columns: front<TAB>back[<TAB>tags]  — lines starting with # are ignored
    For JSON:
      Array of {front, back, tags} objects or [front, back, tags] arrays
    For APKG:
      Standard Anki package; front/back resolved from model field names

    tags: semicolon-separated (space-separated accepted for TSV/JSON/APKG).
    """

    permission_classes = [IsManagementUser]
    parser_classes = [MultiPartParser]

    def post(self, request):  # type: ignore[no-untyped-def]
        import_file = request.FILES.get("import_file") or request.FILES.get("csv_file")
        if not import_file:
            return Response({"detail": "import_file is required."}, status=status.HTTP_400_BAD_REQUEST)

        filename = import_file.name or ""
        ext = filename.rsplit(".", 1)[-1] if "." in filename else "csv"

        deck_id = request.data.get("deck_id")
        deck_name = (request.data.get("deck_name") or "").strip() or "Imported"
        deck_type = (request.data.get("deck_type") or Deck.DeckType.CUSTOM).strip() or Deck.DeckType.CUSTOM

        if deck_type not in {c for c, _ in Deck.DeckType.choices}:
            return Response({"detail": "Invalid deck_type."}, status=400)

        if deck_id:
            try:
                deck = Deck.objects.get(id=deck_id, user=request.user)
            except Deck.DoesNotExist:
                return Response({"detail": "Deck not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            deck, _ = Deck.objects.get_or_create(user=request.user, name=deck_name, defaults={"deck_type": deck_type})

        if deck.is_locked:
            return Response({"detail": "This deck is locked."}, status=status.HTTP_403_FORBIDDEN)

        rows, err = _rows_from_file(import_file, ext)
        if err:
            return Response({"detail": err}, status=400)
        if not rows:
            return Response({"detail": "File contained no importable rows."}, status=400)

        # Validate that at least one of the required field columns is present (CSV/XLSX only)
        if ext in ("csv", "xlsx") and rows:
            keys = set(rows[0].keys())
            if not (keys & {"front", "kanji_character", "vocab_word"}):
                return Response(
                    {"detail": "File must include a 'front' column, or 'kanji_character'/'vocab_word' columns."},
                    status=400,
                )

        created = 0
        skipped = 0

        with transaction.atomic():
            for raw in rows:
                tags_raw = raw.get("tags", "")
                tags = [t.strip() for t in tags_raw.split(";") if t.strip()] if tags_raw else []

                ch = raw.get("kanji_character", "")
                word = raw.get("vocab_word", "")
                reading = raw.get("vocab_reading", "")

                kanji = Kanji.objects.filter(character=ch).first() if ch else None
                vocab = None
                if word:
                    vocab_qs = Vocabulary.objects.filter(word=word)
                    vocab = vocab_qs.filter(reading=reading).first() if reading else vocab_qs.first()

                front = raw.get("front", "")
                back = raw.get("back", "")
                furigana = raw.get("furigana", "")
                image = raw.get("image", "") or raw.get("image_url", "")
                audio = raw.get("audio", "") or raw.get("audio_url", "")

                if not front and kanji:
                    front = kanji.character
                    back = f"{kanji.meaning_en}\nonyomi: {kanji.onyomi}\nkunyomi: {kanji.kunyomi}".strip()
                    if not furigana:
                        furigana = kanji.onyomi
                if not front and vocab:
                    front = vocab.word
                    back = f"{vocab.meaning_en}\nreading: {vocab.reading}".strip()
                    if not furigana:
                        furigana = vocab.reading

                if not front or not back:
                    skipped += 1
                    continue

                Card.objects.create(
                    deck=deck, kanji=kanji, vocab=vocab,
                    front=front, back=back, furigana=furigana,
                    image=image, audio=audio, tags=tags,
                )
                created += 1

        ImportLog.objects.create(
            user=request.user,
            content_type=ImportLog.ContentType.FLASHCARD,
            filename=import_file.name,
            file_format=ext,
            rows_imported=created,
            rows_skipped=skipped,
            extra={"deck_id": deck.id, "deck_name": deck.name},
        )
        return Response({"deck_id": deck.id, "created": created, "skipped": skipped}, status=status.HTTP_201_CREATED)


class ImportLogListView(APIView):
    """GET  /api/flash/import-log/      — paginated list of import history (newest first)
    DELETE /api/flash/import-log/<id>/  — delete a log entry
    """

    permission_classes = [IsManagementUser]

    def get(self, request):  # type: ignore[no-untyped-def]
        qs = ImportLog.objects.filter(user=request.user)
        serializer = ImportLogSerializer(qs[:100], many=True)
        return Response({"results": serializer.data, "count": qs.count()})

    def delete(self, request, log_id: int):  # type: ignore[no-untyped-def]
        try:
            log = ImportLog.objects.get(id=log_id, user=request.user)
            log.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ImportLog.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
